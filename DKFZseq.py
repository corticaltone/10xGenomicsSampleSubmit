#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Apr 25 10:19:07 2019
DKFZseq.py will prepare a sample submission form for sequencing
10xgenomics libraries. Input file should have sample names and 
index well identifiers

@author: SciCom
"""

import json
from openpyxl import Workbook
from openpyxl import load_workbook

# load sample files
ss = load_workbook('SampleTemplateMultiplex.xlsx')
ws = ss.active
sampleCol = ws['A']
wellCol = ws['B']
sampleIndecies = {}
for samples in range(1, len(sampleCol)):
    sampleIndecies[sampleCol[samples].value] = wellCol[samples].value
print(sampleIndecies)

#Load Chromium indecises and look up those corresponding to samples
with open('chromium-dna-sample-indexes-plate.json') as json_file:  
    data = json.load(json_file)
    wellLetterdict = {
            'A':0,
            'B':1,
            'C':2,
            'D':3,
            'E':4,
            'F':5,
            'G':6,
            'H':7            
            }
    
demultSamples = {}
for samples in sampleIndecies:
    currentWell = sampleIndecies[samples]
    letterComponent = currentWell[0]
    numberComponent = currentWell[1:]

    
    sampleLetterdict = {
            0:'A',
            1:'B',
            2:'C',
            3:'D'
            }
    for i in range (4):
        nextIndex = sampleIndecies[samples]
        letter = wellLetterdict[letterComponent]
        number = int(numberComponent)
        wellNum = 12*letter + number
        currentSample = samples + sampleLetterdict[i]
        demultSamples[currentSample] = data[wellNum][1][i]



#Compile output spreadsheet

outPut = load_workbook('IlseTemplateMultiplex-2.xlsx')
opWS = outPut.active
opWS.delete_rows(2)
for currentRow in ws.iter_rows(min_row=2, values_only=True):
    rowList = currentRow
    for i in range(4):
        opWS.append(rowList)
        


for i in range(len(demultSamples)):
    print(opWS['A'][i].value)
    key = list(demultSamples)[i]
    opWS['A'][i+1].value = key
    opWS['B'][i+1].value = demultSamples[key]
    
outPut.save('/Users/tony/upload.xlsx')

